# stock_updater.py - Pure yfinance technical screening script
# No FMP/Tiingo, full tickers.csv, 13 columns A-M, swapped D/E in summaries

import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

def calculate_category(prices, index):
    """Calculate category based on moving averages and price at a given index."""
    adjusted_index = len(prices) + index if index < 0 else index
    if adjusted_index < 0 or adjusted_index >= len(prices):
        return 'Unknown'
    
    ma_series = prices.rolling(150).mean()
    ma150 = ma_series.iloc[adjusted_index]
    valid_offset = min(42, adjusted_index)
    ma150_two_mo_ago = ma_series.iloc[adjusted_index - valid_offset] if valid_offset > 0 else pd.NA
    price = prices.iloc[adjusted_index]
    
    if pd.isna(ma150_two_mo_ago) or pd.isna(ma150) or pd.isna(price):
        return 'Unknown'
    
    cond_a = ma150 > ma150_two_mo_ago
    cond_b = price > ma150
    
    if cond_a and cond_b:
        return 'Uptrend'
    elif not cond_a and cond_b:
        return 'Snapback'
    elif cond_a and not cond_b:
        return 'Pullback'
    else:
        return 'Downtrend'

def calculate_relative_strength(prices):
    """Calculate relative strength as 12-month return percentage."""
    if len(prices) < 252:
        return pd.NA
    price_12mo_ago = prices.iloc[-252]
    current_price = prices.iloc[-1]
    return_12mo = (current_price / price_12mo_ago - 1) * 100
    result = return_12mo
    if ticker in ['SRPT', 'MYGN']:
        print(f"Debug - {ticker}: 12mo Price={price_12mo_ago}, Current Price={current_price}, 12mo Return={return_12mo:.2f}%, 12-1M Return={result:.2f}%")
    return max(result, -100.0)

# Step 1: Load tickers
try:
    tickers_df = pd.read_csv('/Users/vernonbice/StockAnalysis/tickers.csv')
    if 'Symbol' not in tickers_df.columns:
        raise ValueError("CSV must contain a 'Symbol' column")
    tickers = tickers_df['Symbol'].tolist()  # Full list
    print(f"Loaded {len(tickers)} tickers from tickers.csv")
except Exception as e:
    print(f"Error loading tickers.csv: {e}")
    tickers = ['AAPL', 'MSFT', 'TSLA', 'NVDA']  # fallback

# Step 2: Fetch price data
end_date = datetime.today().strftime('%Y-%m-%d')
start_date = (datetime.today() - timedelta(days=400)).strftime('%Y-%m-%d')
print("Downloading price history from yfinance...")
data = yf.download(tickers, start=start_date, end=end_date, auto_adjust=False, threads=True)['Close']

skipped_stocks = []
categories = []
relative_strength_values = []

for ticker in tickers:
    try:
        prices = data[ticker].dropna()
        if len(prices) < 252:
            skipped_stocks.append({'Ticker': ticker, 'Reason': f'Not enough price data ({len(prices)} days)'})
            continue
        
        ma = prices.rolling(150).mean()
        ma150_today = ma.iloc[-1]
        ma150_two_mo_ago = ma.iloc[-42] if len(prices) > 42 else pd.NA
        price_today = prices.iloc[-1]
        
        category = calculate_category(prices, -1)
        category_one_week = calculate_category(prices, -5)
        category_changed = 'x' if category != category_one_week else ''
        
        if ticker == 'MSFT':
            print(f"MSFT: Price={price_today:.2f}, MA150_Today={ma150_today:.2f}, MA150_2mo_Ago={ma150_two_mo_ago:.2f}, Category={category}")
        
        print(f"Processing data for {ticker}")
        try:
            ticker_info = yf.Ticker(ticker).info
            sector = ticker_info.get('sector', 'Unknown')
            industry = ticker_info.get('industry', 'Unknown')
            
            relative_strength = calculate_relative_strength(prices)
            
            categories.append({
                'Ticker': ticker,
                'Company Name': ticker_info.get('longName', 'Unknown'),
                'Sector': sector,
                'Industry': industry,
                'Trend': category if category != 'Unknown' else 'N/A',
                'Relative Strength': relative_strength,
                '12-1M Return': relative_strength,
                'Market Cap (M)': int(round(ticker_info.get('marketCap', 0) / 1e6)),
                'Price': price_today,
                'MA150_Today': ma150_today,
                'MA150_2mo_Ago': ma150_two_mo_ago,
                'Trend 1 Week Ago': category_one_week if category_one_week != 'Unknown' else 'N/A',
                'Trend Changed': category_changed
            })
            relative_strength_values.append(relative_strength)
        except Exception as e:
            print(f"Error processing info for {ticker}: {e}")
            skipped_stocks.append({'Ticker': ticker, 'Reason': str(e)})
        
    except Exception as e:
        skipped_stocks.append({'Ticker': ticker, 'Reason': str(e)})
        continue

# Validate processed tickers
processed_count = len(categories)
expected = len(tickers) - len(skipped_stocks)
if processed_count != expected:
    print(f"Warning: Processed {processed_count} tickers, expected {expected}")

# Rank Relative Strength
if relative_strength_values:
    rs_df = pd.DataFrame(relative_strength_values, columns=['RS'])
    rs_df['Rank'] = rs_df['RS'].rank(pct=True) * 98 + 1
    rs_df['Rank'] = rs_df['Rank'].round(0).astype(int)
    relative_strength_ranks = rs_df['Rank'].tolist()
    for i, rs in enumerate(relative_strength_ranks):
        categories[i]['Relative Strength'] = rs

# Create DataFrames
df = pd.DataFrame(categories)
if not df.empty:
    df = df.sort_values(by=['Sector', 'Industry', 'Ticker'], na_position='last')

df_skipped = pd.DataFrame(skipped_stocks)

# Create summary tables
if not df.empty:
    df_summary = df[df['Trend'] != 'N/A']
    if not df_summary.empty:
        total_counts = df_summary['Trend'].value_counts()
        total_stocks = len(df_summary)
        total_percentages = (total_counts / total_stocks * 100).round(0).astype(int)
        total_row = pd.DataFrame([total_percentages], columns=['Uptrend', 'Snapback', 'Pullback', 'Downtrend'], index=['Watch List Total'])
        
        sector_data = pd.crosstab(df_summary['Sector'], df_summary['Trend'], normalize='index') * 100
        sector_counts = sector_data.reindex(columns=['Uptrend', 'Snapback', 'Pullback', 'Downtrend'], fill_value=0).round(0).astype(int).sort_index()
        sector_counts = pd.concat([total_row, sector_counts])
        
        industry_data = pd.crosstab(df_summary['Industry'], df_summary['Trend'], normalize='index') * 100
        industry_summary = industry_data.reindex(columns=['Uptrend', 'Snapback', 'Pullback', 'Downtrend'], fill_value=0).round(0).astype(int).sort_index()
    else:
        sector_counts = pd.DataFrame(columns=['Uptrend', 'Snapback', 'Pullback', 'Downtrend'])
        industry_summary = pd.DataFrame(columns=['Uptrend', 'Snapback', 'Pullback', 'Downtrend'])
else:
    sector_counts = pd.DataFrame(columns=['Uptrend', 'Snapback', 'Pullback', 'Downtrend'])
    industry_summary = pd.DataFrame(columns=['Uptrend', 'Snapback', 'Pullback', 'Downtrend'])

# Save to Excel
with pd.ExcelWriter('/Users/vernonbice/StockAnalysis/stock_report.xlsx', engine='openpyxl') as writer:
    if not df.empty:
        df.to_excel(writer, sheet_name='Watch List', index=False)
    if not sector_counts.empty:
        sector_counts.to_excel(writer, sheet_name='Sector Summary')
    if not industry_summary.empty:
        industry_summary.to_excel(writer, sheet_name='Industry Summary')
    if not df_skipped.empty:
        df_skipped.to_excel(writer, sheet_name='Skipped Stocks', index=False)
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for sheet_name in ['Watch List', 'Sector Summary', 'Industry Summary', 'Skipped Stocks']:
        if sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            if sheet_name == 'Watch List':
                worksheet.column_dimensions['B'].width = 35
                worksheet.column_dimensions['C'].width = 20
                worksheet.column_dimensions['D'].width = 31
                worksheet.column_dimensions['E'].width = 12
                worksheet.column_dimensions['F'].width = 19
                worksheet.column_dimensions['G'].width = 16
                worksheet.column_dimensions['H'].width = 17
                worksheet.column_dimensions['I'].width = 11
                worksheet.column_dimensions['J'].width = 16
                worksheet.column_dimensions['K'].width = 19
                worksheet.column_dimensions['L'].width = 19
                worksheet.column_dimensions['M'].width = 17
                worksheet.freeze_panes = 'A2'
            elif sheet_name in ['Sector Summary', 'Industry Summary']:
                worksheet.column_dimensions['A'].width = 35
                for col in ['B', 'C', 'D', 'E']:
                    worksheet.column_dimensions[col].width = 14
            elif sheet_name == 'Skipped Stocks':
                worksheet.column_dimensions['A'].width = 8
                worksheet.column_dimensions['B'].width = 23
            for row in worksheet.rows:
                for cell in row:
                    cell.border = border
                    if isinstance(cell.value, (int, float)) and cell.column_letter != 'A':
                        cell.number_format = '#,##0'
            if sheet_name == 'Watch List':
                for col in ['A', 'B', 'C', 'D']:
                    for cell in worksheet[col]:
                        cell.alignment = Alignment(horizontal='left')
                for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                    for cell in worksheet[col]:
                        cell.alignment = Alignment(horizontal='center')
            else:
                for cell in worksheet['A']:
                    cell.alignment = Alignment(horizontal='left')
                for col in ['B', 'C', 'D', 'E']:
                    for cell in worksheet[col]:
                        cell.alignment = Alignment(horizontal='center')
            worksheet.auto_filter.ref = worksheet.dimensions
    
    if 'Sector Summary' in writer.sheets and worksheet.max_row > 1:
        worksheet = writer.sheets['Sector Summary']
        max_row = worksheet.max_row
        worksheet.cell(row=max_row + 1, column=1).value = f"Footnote: {len(skipped_stocks)} stocks skipped due to insufficient data (listed in Skipped Stocks tab)"
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Sector by Technical Stage"
        chart.y_axis.title = "Sector"
        chart.x_axis.title = "Percentage"
        chart.y_axis.majorUnit = 1
        chart.y_axis.minimum = 0
        chart.y_axis.maximum = max_row - 1
        chart.x_axis.majorUnit = 10
        chart.x_axis.minimum = 0
        chart.x_axis.maximum = 100
        data = Reference(worksheet, min_col=2, max_col=5, min_row=1, max_row=max_row)
        cats = Reference(worksheet, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        series = chart.series
        if len(series) >= 4:
            series[0].graphicalProperties.solidFill = "228B22"  # Uptrend
            series[1].graphicalProperties.solidFill = "4682B4"  # Snapback
            series[2].graphicalProperties.solidFill = "FFD700"  # Pullback
            series[3].graphicalProperties.solidFill = "B22222"  # Downtrend
        chart.y_axis.majorGridlines = None
        chart.dataLabels = None
        chart.gapWidth = 20
        chart.height = 15
        chart.width = 30
        worksheet.add_chart(chart, f"A{max_row + 5}")

print("Saved stock_report.xlsx with tabs: Watch List, Sector Summary, Industry Summary, Skipped Stocks")
print("Report complete! Open stock_report.xlsx in Excel to view all tabs")